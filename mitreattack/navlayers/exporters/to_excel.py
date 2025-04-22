"""Contains ToExcel class."""

from copy import deepcopy
from openpyxl.comments import Comment
from openpyxl.styles import PatternFill, Font
import colorsys

from mitreattack.navlayers.core import Layer, Gradient
from mitreattack.navlayers.exporters import ExcelTemplates


class ToExcel:
    """Class that assists in converting data to Excel files."""

    def __init__(self, domain="enterprise", source="local", resource=None):
        """Set up exporting system, builds underlying matrix.

        :param source: Source to generate the matrix from, one of (local or remote)
        :param resource: Optional string path to local cache of stix data (local) or url of workbench to reach out
                            to (remote)
        """
        if domain in ["enterprise-attack", "mitre-enterprise"]:
            domain = "enterprise"
        elif domain in ["mobile-attack", "mitre-mobile"]:
            domain = "mobile"
        elif domain == "ics-attack":
            domain = "ics"
        self.domain = domain
        self.raw_handle = ExcelTemplates(domain=domain, source=source, resource=resource)

    def to_xlsx(self, layerInit, filepath="layer.xlsx"):
        """Export a layer file to the excel format as the file specified.

        :param layerInit: A layer to initialize the instance with
        :param filepath: The location to write the excel file to
        """
        if not isinstance(layerInit, Layer):
            raise TypeError

        layer = deepcopy(layerInit)

        if self.domain not in layer.layer.domain:
            raise ValueError(f"layer domain ({layer.layer.domain}) does not match exporter domain ({self.domain})")

        included_subs = []
        if layer.layer.techniques:
            for entry in layer.layer.techniques:
                if entry.showSubtechniques:
                    if entry.tactic:
                        included_subs.append((entry.techniqueID, entry.tactic))
                    else:
                        included_subs.append((entry.techniqueID, False))

        excluded = []
        if layer.layer.hideDisabled:
            for entry in layer.layer.techniques:
                if entry.enabled is False:
                    if entry.tactic:
                        excluded.append((entry.techniqueID, entry.tactic))
                    else:
                        excluded.append((entry.techniqueID, False))
        scores = []
        if layer.layer.techniques:
            for entry in layer.layer.techniques:
                if entry.score is not None:
                    if entry.tactic:
                        scores.append((entry.techniqueID, entry.tactic, entry.score))
                    else:
                        scores.append((entry.techniqueID, False, entry.score))
        sName = True
        sID = False
        sort = 0
        if layer.layer.layout:
            sName = layer.layer.layout.showName
            sID = layer.layer.layout.showID
        if layer.layer.sorting:
            sort = layer.layer.sorting
        raw_template = self.raw_handle.export(
            showName=sName,
            showID=sID,
            filters=layer.layer.filters,
            sort=sort,
            scores=scores,
            subtechs=included_subs,
            exclude=excluded,
        )
        sheet_obj = raw_template.active
        sheet_obj.title = layer.layer.name
        # v4.2 - do aggregate adjustments
        if layer.layer.layout:
            if layer.layer.layout.showAggregateScores:
                for tac_column in self.raw_handle.codex:
                    short_hand = self.raw_handle.h.convert(tac_column.tactic.name)
                    for x in tac_column.techniques:
                        x_score = [y for y in scores if (y[0] == x.id and (y[1] == short_hand or y[1] is None))]
                        if len(x_score):
                            x.score = x_score[0][2]
                        subs = tac_column.subtechniques.get(x.id, [])
                        for sub_score in subs:
                            subtech_score = [
                                y for y in scores if (y[0] == sub_score.id and (y[1] == short_hand or y[1] is None))
                            ]
                            if len(subtech_score):
                                sub_score.score = subtech_score[0][2]
                        mod = layer.layer.layout.compute_aggregate(x, subs)
                        patch_target = [
                            y
                            for y in layer.layer.techniques
                            if (y.techniqueID == x.id and (y.tactic == short_hand or y.tactic is None))
                        ]
                        if len(patch_target):
                            patch_target[0].score = mod
                        elif mod:
                            print("[WARNING] - Aggregate calculated for a technique that doesn't seem to exist...")

        # verify gradient information
        safe_gradient = layer.layer.gradient
        if not safe_gradient:
            safe_gradient = Gradient(colors=["#ff6666", "#ffe766", "#8ec843"], minValue=1, maxValue=100)

        for tech in layer.layer.techniques:
            p_tactic = None
            if tech.tactic:
                p_tactic = tech.tactic
            coords = self.raw_handle.retrieve_coords(tech.techniqueID, p_tactic)
            if coords == [] or coords == "HIDDEN":
                tac = p_tactic
                if tac is None:
                    tac = "(none)"
                if coords:
                    print(
                        "WARNING! Technique/Tactic "
                        + tech.techniqueID
                        + "/"
                        + tac
                        + " does not appear to exist in the loaded matrix. Skipping..."
                    )
                else:
                    parents = [x for x in layer.layer.techniques if x.techniqueID == tech.techniqueID.split(".")[0]]
                    if tech.tactic:
                        parents = [x for x in parents if x.tactic == tech.tactic]
                    if all([True if not x.showSubtechniques else False for x in parents]):
                        print(
                            "NOTE! Technique/Tactic " + tech.techniqueID + "/" + tac + " does not appear "
                            "to be visible in the matrix. Its parent appears to be hiding it."
                        )
                    else:
                        print(
                            "WARNING! Technique/Tactic " + tech.techniqueID + "/" + tac + " seems malformed. "
                            "Skipping..."
                        )
                    continue
            for location in coords:
                cell = sheet_obj.cell(row=location[0], column=location[1])
                if tech.comment:
                    cell.comment = Comment(tech.comment, "ATT&CK Scripts Exporter")

                if tech.enabled is False:
                    if layer.layer.hideDisabled:
                        pass
                    else:
                        grayed_out = Font(color="909090")
                        cell.font = grayed_out
                        continue
                if tech.color:
                    c_color = PatternFill(fill_type="solid", start_color=tech.color.upper()[1:])
                    cell.fill = c_color
                    continue
                if tech.score is not None:
                    tscore = tech.score
                    comp_color = safe_gradient.compute_color(tscore)
                    c_color = PatternFill(fill_type="solid", start_color=comp_color.upper()[1:])
                    cell.fill = c_color
                    RGB = tuple(int(comp_color.upper()[1:][i : i + 2], 16) for i in (0, 2, 4))
                    hls = colorsys.rgb_to_hls(RGB[0], RGB[1], RGB[2])
                    if hls[1] < 127.5:
                        white = Font(color="FFFFFF")
                        cell.font = white
        raw_template.save(filepath)
